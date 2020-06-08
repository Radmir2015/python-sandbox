from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import datetime
import os.path

from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

driver = None

def startBrowser(driverPack=None):
	if driverPack is None:
		print("Starting browser!")
		global driver
		driver = webdriver.PhantomJS("phantomjs-2.1.1\\bin\\phantomjs.exe", service_args=['--ignore-ssl-errors=true'])
		# driver = webdriver.Chrome()
		driver.set_page_load_timeout(120)
		print("Browser started!")
	else:
		print("Browser is already running!")
	
def getHtml(url, presentElem=None):
	timeout = 120
	if presentElem:
		try:
			driver.get(url)
			element_present = EC.presence_of_element_located((By.CSS_SELECTOR, presentElem))
			WebDriverWait(driver, timeout).until(element_present)
			return driver.page_source
		except TimeoutException:
			print("Timed out waiting for page to load")
	else:
		try:
			driver.get(url)
			return driver.page_source
		except:
			print("Need for restart browser!")
			driver.quit()
			startBrowser()

def getNumberOfPages(soup):
	return len(soup.findAll("a", "schema-pagination__pages-link"))

numberOfPages = None
counter = 0

def getAllGoodsFromPageOnliner(url, curr, needNumberOfPages=False):
	# soup = BeautifulSoup(getHtml(url, "span[data-bind=\"html: product.extended_name || product.full_name\""), "html.parser")
	soup = BeautifulSoup(getHtml(url, "div.schema-product__group"), "html.parser")
	# soup = BeautifulSoup(getHtml(url), "html.parser")
	containers = soup.findAll("div", {"class": "schema-product__group"})

	global numberOfPages
	global counter

	if needNumberOfPages or not numberOfPages:
		numberOfPages = len(soup.findAll("a", "schema-pagination__pages-link"))

	# if needNumberOfPages:
	# 	numberOfPages = getNumberOfPages(soup)

	goods = [] # [name, price, url]
	
	for c in containers:
		counter += 1
		n = c.find("span", {"data-bind": "html: product.extended_name || product.full_name"})
		price = c.find("span", {"data-bind": "html: $root.format.minPrice($data.prices, 'BYN')"})
		avalible = c.find("div", {"class": "schema-product__status"})
		if avalible:
			goods.append([counter, n.text, avalible.text, n.parent["href"]])
		else:
			price = price.text.replace(",", ".")[:-2].encode("ascii", "ignore")
			price = float(price) / curr["BYN"]
			goods.append([counter, n.text, price, n.parent["href"]])
		# print("Good's name: " + good_name)
		# print("Good's url: " + good_url)
		# print("Good's price: " + good_price)
	return goods

def getAllGoodsFromPageCeneo(url, curr, needNumberOfPages=False):
	try:
		soup = BeautifulSoup(getHtml(url), "html.parser")
	except:
		return False
	containers = soup.findAll("div", {"class": "cat-prod-row js_category-list-item js_man-track-event"})

	global numberOfPages
	global counter

	if needNumberOfPages or not numberOfPages:
		numberOfPages = int(soup.find("input", {"id": "page-counter"})["data-pagecount"])

	goods = []

	for c in containers:
		counter += 1
		link = "https://www.ceneo.pl/" + c["data-pid"]
		name = c.find("strong", {"class": "cat-prod-row-name"}).find("a").text
		price = c.find("span", {"class": "price"})
		price = price.find("span", "value").text + price.find("span", "penny").text.replace(",", ".")
		price = float(price.encode("ascii", "ignore")) / curr["PLN"]
		goods.append([counter, name, price, link])

	try:
		nextUrl = "https://www.ceneo.pl" + soup.find("li", "page-arrow arrow-next").find("a")["href"]
	except:
		nextUrl = ""

	return goods, nextUrl

def saveTotalInExcel(total, fileName):
	wb = Workbook()
	ws = wb.active

	ws.append(["Order", "Good's name", "Price", "URL"])
	for pages in total:
		for page in pages:
			ws.append(page)

	wb.save(fileName)
	print()
	print("Saved in " + fileName)
	print()

def searchGoods(url, choice, curr, fileName):

	startBrowser(driver)

	total = []

	time = datetime.datetime.now()

	if choice == "ceneo":
		url = getAllGoodsFromPageCeneo(url, curr, True)
		total.append(url[0])
		url = url[1]
	elif choice == "onliner":
		total.append(getAllGoodsFromPageOnliner(url, curr, True))

	print("1 / " + str(numberOfPages) + " page done! Spent time: " + str(datetime.datetime.now() - time))

	for n in range(2, numberOfPages + 1):
		time = datetime.datetime.now()

		if choice == "ceneo":
			url = getAllGoodsFromPageCeneo(url, curr)
			if not url:
				break
			total.append(url[0])
			url = url[1]
		elif choice == "onliner":
			total.append(getAllGoodsFromPageOnliner(url + "?page=" + str(n), curr))

		print(str(n) + " / " + str(numberOfPages) + " page done! Spent time: " + str(datetime.datetime.now() - time))

	global counter
	counter = 0

	saveTotalInExcel(total, fileName)

def IntersectionAndSave(onlinerFilename, ceneoFilename, finishFilename):
	goods = []
	goods2 = []

	namesLower = []
	names2Lower = []
	names = []
	names2 = []

	try:
		wb = load_workbook(onlinerFilename)
		ws = wb.active

		count = list(e.value for e in tuple(ws.columns)[0])[1:]
		names = list(e.value for e in tuple(ws.columns)[1])[1:]
		prices = list(e.value for e in tuple(ws.columns)[2])[1:]
		urls = list(e.value for e in tuple(ws.columns)[3])[1:]


		for item in range(len(names)):
			if names[item]:
				goods.append([count[item], names[item], prices[item], urls[item], names[item].lower()])
				namesLower.append(names[item].lower())
			
		# namesLower = list(n.lower() for n in names)
	except:
		print("Onliner base not found ({})".format(onlinerFilename))

	try:
		wb2 = load_workbook(ceneoFilename)
		ws2 = wb2.active
		
		count2 = list(e.value for e in tuple(ws2.columns)[0])[1:]
		names2 = list(e.value for e in tuple(ws2.columns)[1])[1:]
		prices2 = list(e.value for e in tuple(ws2.columns)[2])[1:]
		urls2 = list(e.value for e in tuple(ws2.columns)[3])[1:]

		for item in range(len(names2)):
			if names2[item]:
				goods2.append([count2[item], names2[item], prices2[item], urls2[item], names2[item].lower()])
				names2Lower.append(names2[item].lower())

		# names2Lower = list(n.lower() for n in names2)
	except:
		print("Ceneo base not found ({})".format(ceneoFilename))


	inter = list(set.intersection(set(namesLower), set(names2Lower)))

	print("{} goods in onlinerDatebase and {} goods in ceneoDatabase".format(len(namesLower), len(names2Lower)))
	print(str(len(inter)) + " coinciding goods from both bases")
	# print(set.intersection(set(namesLower), set(names2Lower)))


	if os.path.isfile(finishFilename):
		wbEnd = load_workbook(finishFilename)
		wsEnd = wbEnd.active

		rows = tuple(wsEnd.rows)[1:]

		header = [t.value for t in tuple(wsEnd.rows)[0]]

		interFinish = []
		ceneoFinish = []
		onlinerFinish = []
		temp = []

		global updatedGoods
		updatedGoods = 0
		global addedGoods
		addedGoods = 0
		global mergedGoods
		mergedGoods = 0

		for r in range(len(rows)):
			temp = [cell.value for cell in rows[r]]
			if temp[0] and temp[1]:
				interFinish.append(temp)
			elif temp[0]:
				onlinerFinish.append(temp)
			else:
				ceneoFinish.append(temp)

		def searchInFinish(interFinish, onlinerFinish, ceneoFinish, goodName):
			goodName = goodName.lower()
			try:
				return "inter", [i[2].lower() for i in interFinish].index(goodName)
			except:
				try:
					return "onliner", [i[2].lower() for i in onlinerFinish].index(goodName)
				except:
					try:
						return "ceneo", [i[2].lower() for i in ceneoFinish].index(goodName)
					except:
						return None

		for i in inter:
			searchResult = searchInFinish(interFinish, onlinerFinish, ceneoFinish, i)

			if searchResult and searchResult[0] == "inter":
				goodOnliner = goods[namesLower.index(i)]
				goodCeneo = goods2[names2Lower.index(i)]
				if (goodOnliner[2] != interFinish[searchResult[1]][3]) or (goodCeneo[2] != interFinish[searchResult[1]][4]) or (interFinish[searchResult[1]][0] != goodOnliner[0]) or (interFinish[searchResult[1]][5] != goodOnliner[3]) or (interFinish[searchResult[1]][1] != goodCeneo[0]) or (interFinish[searchResult[1]][6] != goodCeneo[3]):
					updatedGoods += sum(list(map(lambda x: 1 if x else 0, [goodOnliner[2] != interFinish[searchResult[1]][3], goodCeneo[2] != interFinish[searchResult[1]][4]])))
					interFinish[searchResult[1]][3] = goodOnliner[2]
					interFinish[searchResult[1]][4] = goodCeneo[2]
					interFinish[searchResult[1]][0] = goodOnliner[0] # position
					interFinish[searchResult[1]][5] = goodOnliner[3] # url
					interFinish[searchResult[1]][1] = goodCeneo[0] # position
					interFinish[searchResult[1]][6] = goodCeneo[3] # url
					interFinish[searchResult[1]][8] = "Updated"
				del goods[namesLower.index(i)]
				del goods2[names2Lower.index(i)]
				# del namesLower[namesLower.index(i)]
				# goods.remove(i)
				namesLower.remove(i)
				# del names2Lower[names2Lower.index(i)]
				# goods2.remove(i)
				names2Lower.remove(i)

			# if searchResult and (searchResult[0] == "onliner" or searchResult[0] == "ceneo"):

			if searchResult and searchResult[0] == "onliner":
				goodCeneo = goods2[names2Lower.index(i)]
				goodOnliner = goods[namesLower.index(i)]
				onl = onlinerFinish[searchResult[1]]
				if onl[3] != goodOnliner[2] or onl[0] != goodOnliner[0] or onl[5] != goodOnliner[3]:
					onlinerFinish[searchResult[1]][3] = goodOnliner[2]
					onlinerFinish[searchResult[1]][0] = goodOnliner[0] # position
					onlinerFinish[searchResult[1]][5] = goodOnliner[3] # url
					updatedGoods += 1
				interFinish.append([onl[0], goodCeneo[0], onl[2], onl[3], goodCeneo[2], onl[5], goodCeneo[3], onl[7], "Merged"])
				mergedGoods += 1
				del onlinerFinish[searchResult[1]]
				del goods[namesLower.index(i)]
				# del namesLower[namesLower.index(i)]
				# goods.remove(i)
				namesLower.remove(i)

			if searchResult and searchResult[0] == "ceneo":
				goodOnliner = goods[namesLower.index(i)]
				goodCeneo = goods2[names2Lower.index(i)]
				cen = ceneoFinish[searchResult[1]]
				if cen[4] != goodCeneo[2] or cen[1] != goodCeneo[0] or cen[6] != goodCeneo[3]:
					ceneoFinish[searchResult[1]][4] = goodCeneo[2]
					ceneoFinish[searchResult[1]][1] = goodCeneo[0] # position
					ceneoFinish[searchResult[1]][6] = goodCeneo[3] # url
					updatedGoods += 1
				interFinish.append([goodOnliner[0], cen[1], goodOnliner[1], goodOnliner[2], cen[4], goodOnliner[3], cen[6], cen[7], "Merged"])
				mergedGoods += 1
				del ceneoFinish[searchResult[1]]
				del goods2[names2Lower.index(i)]
				# del names2Lower[names2Lower.index(i)]
				# goods2.remove(i)
				names2Lower.remove(i)

			if searchResult is None:
				goodOnliner = goods[namesLower.index(i)]
				goodCeneo = goods2[names2Lower.index(i)]
				interFinish.append([goodOnliner[0], goodCeneo[0], goodOnliner[1], goodOnliner[2], goodCeneo[2], goodOnliner[3], goodCeneo[3], 0, "Added"])
				addedGoods += 2
				del goods[namesLower.index(i)]
				del goods2[names2Lower.index(i)]
				# del namesLower[namesLower.index(i)]
				# goods.remove(i)
				namesLower.remove(i)
				# goods2.remove(i)
				names2Lower.remove(i)
				# del names2Lower[names2Lower.index(i)]

			# wsEnd.append([goodOnliner[0], goodCeneo[0], goodOnliner[1], goodOnliner[2], goodCeneo[2], goodOnliner[3], goodCeneo[3], 0, "Added"])
		# print(list(g[1] for g in goods))
		# print(list(o[2] for o in onlinerFinish))
		# print(searchInFinish(interFinish, onlinerFinish, ceneoFinish, "Lenovo Legion Y520-15 [80WK002MRK]"))
		goodsCopy = list(goods)
		for g in goodsCopy: # onliner
			searchResult = searchInFinish(interFinish, onlinerFinish, ceneoFinish, g[1])

			if searchResult and searchResult[0] == "inter":
				goodOnliner = goods[namesLower.index(g[4])]
				if interFinish[searchResult[1]][3] != goodOnliner[2] or interFinish[searchResult[1]][0] != goodOnliner[0] or interFinish[searchResult[1]][5] != goodOnliner[3]:
					interFinish[searchResult[1]][3] = goodOnliner[2] # price
					interFinish[searchResult[1]][0] = goodOnliner[0] # position
					interFinish[searchResult[1]][5] = goodOnliner[3] # url
					interFinish[searchResult[1]][8] = "Updated"
					updatedGoods += 1
				goods.remove(g)
				namesLower.remove(g[4])

			if searchResult and searchResult[0] == "onliner":
				goodOnliner = goods[namesLower.index(g[4])]
				if onlinerFinish[searchResult[1]][3] != goodOnliner[2] or onlinerFinish[searchResult[1]][0] != goodOnliner[0] or onlinerFinish[searchResult[1]][5] != goodOnliner[3]:
					onlinerFinish[searchResult[1]][3] = goodOnliner[2]
					onlinerFinish[searchResult[1]][0] = goodOnliner[0] # position
					onlinerFinish[searchResult[1]][5] = goodOnliner[3] # url
					onlinerFinish[searchResult[1]][8] = "Updated"
					updatedGoods += 1
				# print(goods[namesLower.index(g[4])], end=" ")
				goods.remove(g)
				namesLower.remove(g[4])

			if searchResult and searchResult[0] == "ceneo":
				goodOnliner = goods[namesLower.index(g[4])]
				# goodCeneo = goods2[names2Lower.index(g)]
				cen = ceneoFinish[searchResult[1]]
				interFinish.append([goodOnliner[0], cen[1], goodOnliner[1], goodOnliner[2], cen[4], goodOnliner[3], cen[6], cen[7], "Merged"])
				mergedGoods += 1
				del ceneoFinish[searchResult[1]]
				goods.remove(g)
				namesLower.remove(g[4])

			if searchResult is None:
				goodOnliner = goods[namesLower.index(g[4])]
				onlinerFinish.append([goodOnliner[0], "", goodOnliner[1], goodOnliner[2], "", goodOnliner[3], "", 0, "Added"])
				addedGoods += 1

		goodsCopy = list(goods2)
		for g in goodsCopy: # ceneo
			searchResult = searchInFinish(interFinish, onlinerFinish, ceneoFinish, g[1])

			if searchResult and searchResult[0] == "inter":
				goodCeneo = goods2[names2Lower.index(g[4])]
				if interFinish[searchResult[1]][4] != goodCeneo[2] or interFinish[searchResult[1]][1] != goodCeneo[0] or interFinish[searchResult[1]][6] != goodCeneo[3]:
					interFinish[searchResult[1]][4] = goodCeneo[2]
					interFinish[searchResult[1]][1] = goodCeneo[0] # position
					interFinish[searchResult[1]][6] = goodCeneo[3] # url
					interFinish[searchResult[1]][8] = "Updated"
					updatedGoods += 1
				# del goods2[names2Lower.index(g[4])]
				# del names2Lower[names2Lower.index(g[4])]
				goods2.remove(g)
				names2Lower.remove(g[4])

			if searchResult and searchResult[0] == "ceneo":
				goodCeneo = goods2[names2Lower.index(g[4])]
				if ceneoFinish[searchResult[1]][4] != goodCeneo[2] or ceneoFinish[searchResult[1]][1] != goodCeneo[0] or ceneoFinish[searchResult[1]][6] != goodCeneo[3]:
					ceneoFinish[searchResult[1]][4] = goodCeneo[2]
					ceneoFinish[searchResult[1]][1] = goodCeneo[0] # position
					ceneoFinish[searchResult[1]][6] = goodCeneo[3] # url
					ceneoFinish[searchResult[1]][8] = "Updated"
					updatedGoods += 1
				# del goods2[names2Lower.index(g[4])]
				# del names2Lower[names2Lower.index(g[4])]
				goods2.remove(g)
				names2Lower.remove(g[4])

			if searchResult and searchResult[0] == "onliner":
				goodCeneo = goods[namesLower.index(g[4])]
				# goodCeneo = goods2[names2Lower.index(g)]
				onl = onlinerFinish[searchResult[1]]
				interFinish.append([onl[0], goodCeneo[0], onl[2], onl[3], goodCeneo[2], onl[5], goodCeneo[3], onl[7], "Merged"])
				mergedGoods += 1
				del onlinerFinish[searchResult[1]]
				# del goods2[names2Lower.index(g[4])]
				# del names2Lower[names2Lower.index(g[4])]
				goods2.remove(g)
				names2Lower.remove(g[4])

			if searchResult is None:
				goodCeneo = goods2[names2Lower.index(g[4])]
				ceneoFinish.append(["", goodCeneo[0], goodCeneo[1], "", goodCeneo[2], "", goodCeneo[3], 0, "Added"])
				addedGoods += 1

		print("Added goods: ", addedGoods)
		print("Merged goods: ", mergedGoods)
		print("Updated goods: ", updatedGoods)

		# print(onlinerFinish)
		# print(ceneoFinish)

		onlinerFinish.sort(key=lambda x: x[0])
		ceneoFinish.sort(key=lambda x: x[1])

		wbEnd.save(finishFilename)

		wbEnd = Workbook()
		wsEnd = wbEnd.active

		wsEnd.append(header)

		for item in interFinish:
			wsEnd.append(item)

		for item in onlinerFinish:
			wsEnd.append(item)

		for item in ceneoFinish:
			wsEnd.append(item)

		wbEnd.save(finishFilename)


	else:
		wbEnd = Workbook()
		wsEnd = wbEnd.active
		wsEnd.append(["Number Onliner", "Number Ceneo", "Good's name", "Price on Onliner", "Price on Ceneo", "Link on Onliner", "Link on Ceneo", "Need to parse", "Status"])

		for i in inter: # append in table intersections
			# print(namesLower.index(i), names2Lower.index(i))
			goodOnliner = goods[namesLower.index(i)]
			goodCeneo = goods2[names2Lower.index(i)]
			wsEnd.append([goodOnliner[0], goodCeneo[0], goodOnliner[1], goodOnliner[2], goodCeneo[2], goodOnliner[3], goodCeneo[3], 0, "Added"])
			del goods[namesLower.index(i)]
			del namesLower[namesLower.index(i)]
			del goods2[names2Lower.index(i)]
			del names2Lower[names2Lower.index(i)]

		for g in goods: # append in table onliner
			wsEnd.append([g[0], "", g[1], g[2], "", g[3], "", 0, "Added"])

		for g in goods2: #append in table ceneo
			wsEnd.append(["", g[0], g[1], "", g[2], "", g[3], 0, "Added"])

		wbEnd.save(finishFilename)

def updatePricesWithoutThreadings(sheetFilename, curr, time=None):

	print("Opening file: " + sheetFilename)

	wb = load_workbook(sheetFilename)
	sheet = wb.active

	# for s in sheet["H"]:
	# 	print(str(s.value) + " ", end="")

	needUpdate = list(inx for inx, s in enumerate(sheet["H"]) if s.value == 1)
	# print(needUpdate)

	columnsToUpdate = list(tuple(sheet.rows)[n] for n in needUpdate)

	linksAddress = (5, 6)
	pricesAddress = (3, 4)

	if len(columnsToUpdate) > 0:
		print("Need to update: " + str(len(columnsToUpdate)) + " good(s)")
	else:
		print("Update is not necessary")

	for c in range(len(columnsToUpdate)):

		sec = datetime.datetime.now()

		urlOnliner = columnsToUpdate[c][linksAddress[0]].value

		if urlOnliner:
			try:
				html = requests.get(urlOnliner, verify=True, timeout=time).content

				soup = BeautifulSoup(html, "html.parser")

				try:
					onlinerPrice = float(soup.find("span", "product-aside__price--primary").text.strip().replace(",", ".")[:-2]) / curr["BYN"]
				except:
					onlinerPrice = "Нет в наличии"

				columnsToUpdate[c][8].value = "Updated"
				columnsToUpdate[c][7].value = 1 # put 0 to prevent the product from being updated
				columnsToUpdate[c][pricesAddress[0]].value = onlinerPrice

			except:
				# onlinerPrice = "Ошибка получения"
				columnsToUpdate[c][8].value = "Error"
				columnsToUpdate[c][7].value = 1
				return False

		urlCeneo = columnsToUpdate[c][linksAddress[1]].value

		if urlCeneo:
			try:
				html = requests.get(urlCeneo, verify=True, timeout=time).content

				soup = BeautifulSoup(html, "html.parser")

				try:
					ceneoPrice = soup.find("span", {"class": "price"})
					ceneoPrice = float(ceneoPrice.find("span", "value").text + ceneoPrice.find("span", "penny").text.replace(",", ".")) / curr["PLN"]
				except:
					ceneoPrice = "Нет в наличии"

				columnsToUpdate[c][8].value = "Updated"
				columnsToUpdate[c][7].value = 1 # put 0 to prevent the product from being updated
				columnsToUpdate[c][pricesAddress[1]].value = ceneoPrice

			except:
				# ceneoPrice = "Ошибка получения"
				columnsToUpdate[c][8].value = "Error"
				columnsToUpdate[c][7].value = 1
				return False

		print(str(c + 1) + " / " + str(len(columnsToUpdate)) + " good updated! Spent time: " + str(datetime.datetime.now() - sec))

	# print(onlinerPrice, ceneoPrice)

	print("Saving file: " + sheetFilename)

	wb.save(sheetFilename)

def oneThread(columnsToUpdate, c, curr, time=None, proxys=None):

	linksAddress = (5, 6)
	pricesAddress = (3, 4)

	sec = datetime.datetime.now()

	urlOnliner = columnsToUpdate[c][linksAddress[0]].value

	if urlOnliner:
		try:
			html = requests.get(urlOnliner, timeout=time, proxies=proxys).content

			soup = BeautifulSoup(html, "html.parser")

			try:
				onlinerPrice = float(soup.find("span", "product-aside__price--primary").text.strip().replace(",", ".")[:-2]) / curr["BYN"]
			except:
				onlinerPrice = "Нет в наличии"

			columnsToUpdate[c][8].value = "Updated"
			columnsToUpdate[c][7].value = 1 # put 0 to prevent the product from being updated
			columnsToUpdate[c][pricesAddress[0]].value = onlinerPrice

		except:
			# onlinerPrice = "Ошибка получения"
			columnsToUpdate[c][8].value = "Error"
			columnsToUpdate[c][7].value = 1
			return False

	urlCeneo = columnsToUpdate[c][linksAddress[1]].value

	if urlCeneo:
		try:
			html = requests.get(urlCeneo, timeout=time, proxies=proxys).content

			soup = BeautifulSoup(html, "html.parser")

			try:
				ceneoPrice = soup.find("span", {"class": "price"})
				ceneoPrice = float(ceneoPrice.find("span", "value").text + ceneoPrice.find("span", "penny").text.replace(",", ".")) / curr["PLN"]
			except:
				ceneoPrice = "Нет в наличии"

			columnsToUpdate[c][8].value = "Updated"
			columnsToUpdate[c][7].value = 1 # put 0 to prevent the product from being updated
			columnsToUpdate[c][pricesAddress[1]].value = ceneoPrice

		except:
			# ceneoPrice = "Ошибка получения"
			columnsToUpdate[c][8].value = "Error"
			columnsToUpdate[c][7].value = 1
			return False

	print(str(c + 1) + " / " + str(len(columnsToUpdate)) + " good updated! Spent time: " + str(datetime.datetime.now() - sec))

def updatePrices(sheetFilename, curr, time=None, proxys=None):

	from multiprocessing.dummy import Pool as ThreadPool 
	pool = ThreadPool(50)

	print("Opening file: " + sheetFilename)

	wb = load_workbook(sheetFilename)
	sheet = wb.active

	# for s in sheet["H"]:
	# 	print(str(s.value) + " ", end="")

	needUpdate = list(inx for inx, s in enumerate(sheet["H"]) if s.value == 1)
	# print(needUpdate)

	columnsToUpdate = list(tuple(sheet.rows)[n] for n in needUpdate)

	linksAddress = (5, 6)
	pricesAddress = (3, 4)

	if len(columnsToUpdate) > 0:
		print("Need to update: " + str(len(columnsToUpdate)) + " good(s)")
	else:
		print("Update is not necessary")

	# for c in range(len(columnsToUpdate)):
	pool.starmap(oneThread, zip([columnsToUpdate] * len(columnsToUpdate), range(len(columnsToUpdate)), [curr] * len(columnsToUpdate), [time] * len(columnsToUpdate), [proxys] * len(columnsToUpdate)))

	print("Saving file: " + sheetFilename)

	wb.save(sheetFilename)

def getCurrencies(sheetFilename, api_key, base="USD"):
	wb = load_workbook(sheetFilename)
	ws = wb.active

	needUpdate = list(tuple(ws.rows)[i] for i, w in enumerate(ws["C"][1:], 1) if w.value == 1)
	tableCurr = list(tuple(ws.rows)[1:])

	# print(needUpdate)
	USDPLN, USDBYN = 0, 0

	if needUpdate:
		curr = list(n[0].value for n in needUpdate)
		print("Need update this currencies: " + ", ".join(curr))
		currencies = ",".join(list(n[0].value.upper() for n in needUpdate))
		json = requests.get("http://apilayer.net/api/live?access_key=" + api_key + "&currencies=" + currencies + "&source=" + base + "&format=1").json()
		quotes = json["quotes"]
		for n in needUpdate:
			n[1].value = quotes["USD" + n[0].value.upper()]
			n[2].value = 0
			print("USD" + n[0].value.upper() + ": " + str(n[1].value))

	else:
		print("Currency update isn't needed")

	for t in tableCurr:
		if t[0].value == "PLN":
			USDPLN = t[1].value
		if t[0].value == "BYN":
			USDBYN = t[1].value

	wb.save(sheetFilename)
	print("BYN = " + str(USDBYN))
	print("PLN = " + str(USDPLN))
	return {"PLN": USDPLN, "BYN": USDBYN}

def runTask(mainFilename, multiprocessing=False):
	wb = load_workbook(mainFilename)
	ws = wb.active

	priceParse = [] # filenames
	makeData = []  # onliner url, ceneo url, filename

	# print(ws["C"][1])
	taskForParse = list([w, ws["C"][w].value] for w in range(ws.max_row) if ws["C"][w].value != None)[1:]
	# print(taskForParse)
	for column, task in taskForParse:
		if task == 2:
			priceParse.append(ws["D"][column].value)
		if task == 1:
			makeData.append([ws["A"][column].value, ws["B"][column].value, ws["D"][column].value])

	# print(priceParse, makeData)

	file = open("api_key.txt", "r")
	apikey = file.read().strip()
	file.close()

	curr = getCurrencies("currencies.xlsx", apikey)

	for p in priceParse:
		try:
			if multiprocessing:
				updatePrices(p, curr)
			else:
				updatePricesWithoutThreadings(p, curr)
		except:
			print("Файл {} не существует или произошла ошибка при парсинге".format(p))

	for m in makeData:
		filenameOnliner, filenameCeneo = "", ""
		if m[0]:
			filenameOnliner = m[0][len(m[0]) - 1 - m[0][::-1].index("/") + 1:] + "(onliner.by).xlsx"
			searchGoods(m[0], "onliner", curr, filenameOnliner)
		if m[1]:
			filenameCeneo = m[1][len(m[1]) - 1 - m[1][::-1].index("/") + 1:] + "(ceneo.pl).xlsx"
			searchGoods(m[1], "ceneo", curr, filenameCeneo)
		IntersectionAndSave(filenameOnliner, filenameCeneo, m[2])


# print(getAllGoodsFromPageCeneo(page)[1])
# page = "https://catalog.onliner.by/coffee"
# searchGoods(page, "onliner", "onliner.by")
# page = "https://www.ceneo.pl/Ekspresy_do_kawy"
# searchGoods(page, "ceneo", "ceneo.pl")

# IntersectionAndSave("onliner.by", "ceneo.pl", "finishTable")

# updatePricesWithoutThreadings("finishTable")
# proxys = {
# 		"https": "https://139.59.125.12:8080",
# 		"http": "http://89.187.217.115:80"
# 	}
# updatePrices("finishTable", time=10)

if __name__ == "__main__":
	totalTime = datetime.datetime.now()
	runTask("main.xlsx")
	# IntersectionAndSave("notebook(onliner.by).xlsx", "Laptopy(ceneo.pl).xlsx", "notebooks.xlsx")
	print("Totally spent time: " + str(datetime.datetime.now() - totalTime))
	input("Press any key for exit...")