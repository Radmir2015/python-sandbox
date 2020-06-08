from selenium import webdriver
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl import load_workbook
import requests
import datetime
import os.path

from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

driver = None

def startBrowser(driverPack=None):
	if driverPack is None:
		print("Starting browser!")
		global driver
		driver = webdriver.PhantomJS("phantomjs-2.1.1\\bin\\phantomjs.exe", service_args=['--ignore-ssl-errors=true'])
		# driver = webdriver.Chrome()
		driver.set_page_load_timeout(120)
		print("Browser started!")
	else:
		print("Browser is already running!")
	
def getHtml(url, presentElem=None):
	timeout = 120
	if presentElem:
		try:
			driver.get(url)
			element_present = EC.presence_of_element_located((By.CSS_SELECTOR, presentElem))
			WebDriverWait(driver, timeout).until(element_present)
			return driver.page_source
		except TimeoutException:
			print("Timed out waiting for page to load")
	else:
		try:
			driver.get(url)
			return driver.page_source
		except:
			print("Need for restart browser!")
			driver.quit()
			startBrowser()

def getNumberOfPages(soup):
	return len(soup.findAll("a", "schema-pagination__pages-link"))

numberOfPages = None
counter = 0

def getAllGoodsFromPageOnliner(url, curr, needNumberOfPages=False):
	# soup = BeautifulSoup(getHtml(url, "span[data-bind=\"html: product.extended_name || product.full_name\""), "html.parser")
	soup = BeautifulSoup(getHtml(url, "div.schema-product__group"), "html.parser")
	# soup = BeautifulSoup(getHtml(url), "html.parser")
	containers = soup.findAll("div", {"class": "schema-product__group"})

	global numberOfPages
	global counter

	if needNumberOfPages or not numberOfPages:
		numberOfPages = len(soup.findAll("a", "schema-pagination__pages-link"))

	# if needNumberOfPages:
	# 	numberOfPages = getNumberOfPages(soup)

	goods = [] # [name, price, url]
	
	for c in containers:
		counter += 1
		n = c.find("span", {"data-bind": "html: product.extended_name || product.full_name"})
		price = c.find("span", {"data-bind": "html: $root.format.minPrice($data.prices, 'BYN')"})
		avalible = c.find("div", {"class": "schema-product__status"})
		if avalible:
			goods.append([counter, n.text, avalible.text, n.parent["href"]])
		else:
			price = price.text.replace(",", ".")[:-2].encode("ascii", "ignore")
			price = round(float(price) / curr["BYN"], 2)
			goods.append([counter, n.text, price, n.parent["href"]])
		# print("Good's name: " + good_name)
		# print("Good's url: " + good_url)
		# print("Good's price: " + good_price)
	return goods

def getAllGoodsFromPageCeneo(url, curr, needNumberOfPages=False):
	try:
		soup = BeautifulSoup(getHtml(url), "html.parser")
	except:
		return False
	containers = soup.findAll("div", {"class": "cat-prod-row js_category-list-item js_man-track-event"})

	global numberOfPages
	global counter

	if needNumberOfPages or not numberOfPages:
		numberOfPages = int(soup.find("input", {"id": "page-counter"})["data-pagecount"])

	goods = []

	for c in containers:
		counter += 1
		link = "https://ceneo.pl/" + c["data-pid"]
		name = c.find("strong", {"class": "cat-prod-row-name"}).find("a").text
		price = c.find("span", {"class": "price"})
		price = price.find("span", "value").text + price.find("span", "penny").text.replace(",", ".")
		price = round(float(price.encode("ascii", "ignore")) / curr["PLN"], 2)
		goods.append([counter, name, price, link])

	try:
		nextUrl = "https://ceneo.pl" + soup.find("li", "page-arrow arrow-next").find("a")["href"]
	except:
		nextUrl = ""

	return goods, nextUrl

def saveTotalInExcel(total, fileName):
	wb = Workbook()
	ws = wb.active

	ws.append(["Order", "Good's name", "Price", "URL"])
	for pages in total:
		for page in pages:
			ws.append(page)

	wb.save(fileName)
	print()
	print("Saved in " + fileName)
	print()

def searchGoods(url, choice, curr, fileName):

	startBrowser(driver)

	total = []

	time = datetime.datetime.now()

	if choice == "ceneo":
		url = getAllGoodsFromPageCeneo(url, curr, True)
		total.append(url[0])
		url = url[1]
	elif choice == "onliner":
		total.append(getAllGoodsFromPageOnliner(url, curr, True))

	print("1 / " + str(numberOfPages) + " page done! Spent time: " + str(datetime.datetime.now() - time))

	for n in range(2, numberOfPages + 1):
		time = datetime.datetime.now()

		try:

			if choice == "ceneo":
				url = getAllGoodsFromPageCeneo(url, curr)
				if not url:
					break
				total.append(url[0])
				url = url[1]
			elif choice == "onliner":
				total.append(getAllGoodsFromPageOnliner(url + "?page=" + str(n), curr))

			print(str(n) + " / " + str(numberOfPages) + " page done! Spent time: " + str(datetime.datetime.now() - time))

		except:

			print("Parse error on page: " + str(n) + "; continued download...")

	global counter
	counter = 0

	saveTotalInExcel(total, fileName)

def deleteDublecates(sheetFilename):
	wb = load_workbook(sheetFilename)
	ws = wb.active

	header = [t.value for t in tuple(ws.rows)[0]], [t.style for t in tuple(ws.rows)[0]], ws[1]

	goods = list(list(map(lambda x: x.value, t)) for t in tuple(ws.rows)[1:])

	global deletedGoods
	deletedGoods = 0

	# print([goods[i] for i in range(10)])
	# onlinerLinks = [g[5] for g in goods]
	# ceneoLinks = [g[6] for g in goods]

	onlinerLinks, ceneoLinks = [], []

	# print(goods)

	goodsCopy = list(goods)

	for g in goods:
		try:
			if g[5]:
				onlinerLinks.index(g[5])
				goodsCopy.remove(g)
				# print("Removed:", g)
				deletedGoods += 1
		except:
			onlinerLinks.append(g[5])
			try:
				if g[6]:
					ceneoLinks.index(g[6])
					goodsCopy.remove(g)
					# print("Removed:", g)
					deletedGoods += 1
			except:
				ceneoLinks.append(g[6])

	ceneoLinks, onlinerLinks = [], []

	for g in goods:
		try:
			if g[6]:
				ceneoLinks.index(g[6])
				goodsCopy.remove(g)
				# print("Removed:", g)
				deletedGoods += 1
		except:
			ceneoLinks.append(g[6])
			try:
				if g[5]:
					onlinerLinks.index(g[5])
					goodsCopy.remove(g)
					# print("Removed:", g)
					deletedGoods += 1
			except:
				onlinerLinks.append(g[5])

	goods = goodsCopy

	# wb = Workbook()
	# ws = wb.active

	# from copy import copy, deepcopy

	# if header:
	# 	ws.append(header[0])
	# 	# for ind, t in enumerate(tuple(ws[1])):
	# 	# 	t = copy(header[2][ind])
	# 	for ind, t in enumerate(tuple(ws[1])):
	# 		t.style = header[1][ind]

	# for g in goods:
	# 	ws.append(g)

	saveWithHeader(sheetFilename, goods)

	# wb.save(sheetFilename)

def IntersectionAndSave(onlinerFilename, ceneoFilename, finishFilename, way):
	goods = []
	goods2 = []

	namesLower = []
	names2Lower = []
	names = []
	names2 = []

	try:
		wb = load_workbook(onlinerFilename)
		ws = wb.active

		count = list(e.value for e in tuple(ws.columns)[0])[1:]
		names = list(e.value for e in tuple(ws.columns)[1])[1:]
		prices = list(e.value for e in tuple(ws.columns)[2])[1:]
		urls = list(e.value for e in tuple(ws.columns)[3])[1:]


		for item in range(len(names)):
			if names[item]:
				goods.append([count[item], names[item], prices[item], urls[item], names[item].lower()])
				namesLower.append(names[item].lower())
			
		# namesLower = list(n.lower() for n in names)
	except:
		print("Onliner base not found ({})".format(onlinerFilename))

	try:
		wb2 = load_workbook(ceneoFilename)
		ws2 = wb2.active
		
		count2 = list(e.value for e in tuple(ws2.columns)[0])[1:]
		names2 = list(e.value for e in tuple(ws2.columns)[1])[1:]
		prices2 = list(e.value for e in tuple(ws2.columns)[2])[1:]
		urls2 = list(e.value for e in tuple(ws2.columns)[3])[1:]

		for item in range(len(names2)):
			if names2[item]:
				goods2.append([count2[item], names2[item], prices2[item], urls2[item], names2[item].lower()])
				names2Lower.append(names2[item].lower())

		# names2Lower = list(n.lower() for n in names2)
	except:
		print("Ceneo base not found ({})".format(ceneoFilename))


	inter = list(set.intersection(set(namesLower), set(names2Lower)))

	print("{} goods in onlinerDatebase and {} goods in ceneoDatabase".format(len(namesLower), len(names2Lower)))
	print(str(len(inter)) + " coinciding goods from both bases")
	# print(set.intersection(set(namesLower), set(names2Lower)))


	if os.path.isfile(finishFilename):

		deleteDublecates(finishFilename)
		print("Deleted dublecates: " + str(deletedGoods))

		wbEnd = load_workbook(finishFilename)
		wsEnd = wbEnd.active

		rows = tuple(wsEnd.rows)[1:]

		header = [t.value for t in tuple(wsEnd.rows)[0]]

		interFinish = []
		ceneoFinish = []
		onlinerFinish = []
		temp = []

		global updatedGoods
		updatedGoods = 0
		global addedGoods
		addedGoods = 0
		global mergedGoods
		mergedGoods = 0

		for r in range(len(rows)):
			temp = [cell.value for cell in rows[r]]
			if temp[0] and temp[1]:
				interFinish.append(temp)
			elif temp[0]:
				onlinerFinish.append(temp)
			else:
				ceneoFinish.append(temp)

		def searchInFinish(interFinish, onlinerFinish, ceneoFinish, propName, prop):
			prop = prop.lower()
			if propName == "link":
				propName = [5, 6]
			elif propName == "name":
				propName = [2, 2]
			try:
				return "inter", [i[propName[0]].lower() for i in interFinish].index(prop)
			except:
				try:
					return "onliner", [i[propName[1]].lower() for i in onlinerFinish].index(prop)
				except:
					try:
						return "ceneo", [i[propName[1]].lower() for i in ceneoFinish].index(prop)
					except:
						return None

		# for i in inter:
		# 	searchResult = searchInFinish(interFinish, onlinerFinish, ceneoFinish, "link", i)

		# 	if searchResult and searchResult[0] == "inter":
		# 		goodOnliner = goods[namesLower.index(i)]
		# 		goodCeneo = goods2[names2Lower.index(i)]
		# 		if (goodOnliner[2] != interFinish[searchResult[1]][3]) or (goodCeneo[2] != interFinish[searchResult[1]][4]) or (interFinish[searchResult[1]][0] != goodOnliner[0]) or (interFinish[searchResult[1]][5] != goodOnliner[3]) or (interFinish[searchResult[1]][1] != goodCeneo[0]) or (interFinish[searchResult[1]][6] != goodCeneo[3]):
		# 			updatedGoods += sum(list(map(lambda x: 1 if x else 0, [goodOnliner[2] != interFinish[searchResult[1]][3], goodCeneo[2] != interFinish[searchResult[1]][4]])))
		# 			interFinish[searchResult[1]][3] = goodOnliner[2]
		# 			interFinish[searchResult[1]][4] = goodCeneo[2]
		# 			interFinish[searchResult[1]][0] = goodOnliner[0] # position
		# 			interFinish[searchResult[1]][5] = goodOnliner[3] # url
		# 			interFinish[searchResult[1]][1] = goodCeneo[0] # position
		# 			interFinish[searchResult[1]][6] = goodCeneo[3] # url
		# 			interFinish[searchResult[1]][8] = "Updated"
		# 		del goods[namesLower.index(i)]
		# 		del goods2[names2Lower.index(i)]
		# 		# del namesLower[namesLower.index(i)]
		# 		# goods.remove(i)
		# 		namesLower.remove(i)
		# 		# del names2Lower[names2Lower.index(i)]
		# 		# goods2.remove(i)
		# 		names2Lower.remove(i)

		# 	# if searchResult and (searchResult[0] == "onliner" or searchResult[0] == "ceneo"):

		# 	if searchResult and searchResult[0] == "onliner":
		# 		goodCeneo = goods2[names2Lower.index(i)]
		# 		goodOnliner = goods[namesLower.index(i)]
		# 		onl = onlinerFinish[searchResult[1]]
		# 		if onl[3] != goodOnliner[2] or onl[0] != goodOnliner[0] or onl[5] != goodOnliner[3]:
		# 			onlinerFinish[searchResult[1]][3] = goodOnliner[2]
		# 			onlinerFinish[searchResult[1]][0] = goodOnliner[0] # position
		# 			onlinerFinish[searchResult[1]][5] = goodOnliner[3] # url
		# 			updatedGoods += 1
		# 		interFinish.append([onl[0], goodCeneo[0], onl[2], onl[3], goodCeneo[2], onl[5], goodCeneo[3], onl[7], "Merged"])
		# 		mergedGoods += 1
		# 		del onlinerFinish[searchResult[1]]
		# 		del goods[namesLower.index(i)]
		# 		# del namesLower[namesLower.index(i)]
		# 		# goods.remove(i)
		# 		namesLower.remove(i)

		# 	if searchResult and searchResult[0] == "ceneo":
		# 		goodOnliner = goods[namesLower.index(i)]
		# 		goodCeneo = goods2[names2Lower.index(i)]
		# 		cen = ceneoFinish[searchResult[1]]
		# 		if cen[4] != goodCeneo[2] or cen[1] != goodCeneo[0] or cen[6] != goodCeneo[3]:
		# 			ceneoFinish[searchResult[1]][4] = goodCeneo[2]
		# 			ceneoFinish[searchResult[1]][1] = goodCeneo[0] # position
		# 			ceneoFinish[searchResult[1]][6] = goodCeneo[3] # url
		# 			updatedGoods += 1
		# 		interFinish.append([goodOnliner[0], cen[1], goodOnliner[1], goodOnliner[2], cen[4], goodOnliner[3], cen[6], cen[7], "Merged"])
		# 		mergedGoods += 1
		# 		del ceneoFinish[searchResult[1]]
		# 		del goods2[names2Lower.index(i)]
		# 		# del names2Lower[names2Lower.index(i)]
		# 		# goods2.remove(i)
		# 		names2Lower.remove(i)

		# 	if searchResult is None:
		# 		goodOnliner = goods[namesLower.index(i)]
		# 		goodCeneo = goods2[names2Lower.index(i)]
		# 		interFinish.append([goodOnliner[0], goodCeneo[0], goodOnliner[1], goodOnliner[2], goodCeneo[2], goodOnliner[3], goodCeneo[3], 0, "Added"])
		# 		addedGoods += 2
		# 		del goods[namesLower.index(i)]
		# 		del goods2[names2Lower.index(i)]
		# 		# del namesLower[namesLower.index(i)]
		# 		# goods.remove(i)
		# 		namesLower.remove(i)
		# 		# goods2.remove(i)
		# 		names2Lower.remove(i)
		# 		# del names2Lower[names2Lower.index(i)]

		if way == "link":

			goodsCopy = list(goods)
			for g in goodsCopy: # onliner
				searchResult = searchInFinish(interFinish, onlinerFinish, ceneoFinish, "link", g[3])

				if searchResult and searchResult[0] == "inter":
					goodOnliner = goods[namesLower.index(g[4])]
					if interFinish[searchResult[1]][3] != goodOnliner[2] or interFinish[searchResult[1]][0] != goodOnliner[0]: #or interFinish[searchResult[1]][5] != goodOnliner[3]:
						interFinish[searchResult[1]][3] = goodOnliner[2] # price
						interFinish[searchResult[1]][0] = goodOnliner[0] # position
						# interFinish[searchResult[1]][5] = goodOnliner[3] # url
						interFinish[searchResult[1]][8] = "Updated"
						updatedGoods += 1
					goods.remove(g)
					namesLower.remove(g[4])

				if searchResult and searchResult[0] == "onliner":
					goodOnliner = goods[namesLower.index(g[4])]
					if onlinerFinish[searchResult[1]][3] != goodOnliner[2] or onlinerFinish[searchResult[1]][0] != goodOnliner[0]: # or onlinerFinish[searchResult[1]][5] != goodOnliner[3]:
						onlinerFinish[searchResult[1]][3] = goodOnliner[2]
						onlinerFinish[searchResult[1]][0] = goodOnliner[0] # position
						# onlinerFinish[searchResult[1]][5] = goodOnliner[3] # url
						onlinerFinish[searchResult[1]][8] = "Updated"
						updatedGoods += 1
					# print(goods[namesLower.index(g[4])], end=" ")
					goods.remove(g)
					namesLower.remove(g[4])

				if searchResult and searchResult[0] == "ceneo":
					goodOnliner = goods[namesLower.index(g[4])]
					# goodCeneo = goods2[names2Lower.index(g)]
					cen = ceneoFinish[searchResult[1]]
					interFinish.append([goodOnliner[0], cen[1], goodOnliner[1], goodOnliner[2], cen[4], goodOnliner[3], cen[6], cen[7], "Merged"])
					mergedGoods += 1
					del ceneoFinish[searchResult[1]]
					goods.remove(g)
					namesLower.remove(g[4])

				if searchResult is None:
					goodOnliner = goods[namesLower.index(g[4])]
					onlinerFinish.append([goodOnliner[0], "", goodOnliner[1], goodOnliner[2], "", goodOnliner[3], "", 0, "Added"])
					addedGoods += 1

			goodsCopy = list(goods2)
			for g in goodsCopy: # ceneo
				searchResult = searchInFinish(interFinish, onlinerFinish, ceneoFinish, "link", g[3])

				if searchResult and searchResult[0] == "inter":
					goodCeneo = goods2[names2Lower.index(g[4])]
					if interFinish[searchResult[1]][4] != goodCeneo[2] or interFinish[searchResult[1]][1] != goodCeneo[0]: # or interFinish[searchResult[1]][6] != goodCeneo[3]:
						interFinish[searchResult[1]][4] = goodCeneo[2]
						interFinish[searchResult[1]][1] = goodCeneo[0] # position
						# interFinish[searchResult[1]][6] = goodCeneo[3] # url
						interFinish[searchResult[1]][8] = "Updated"
						updatedGoods += 1
					# del goods2[names2Lower.index(g[4])]
					# del names2Lower[names2Lower.index(g[4])]
					goods2.remove(g)
					names2Lower.remove(g[4])

				if searchResult and searchResult[0] == "ceneo":
					goodCeneo = goods2[names2Lower.index(g[4])]
					if ceneoFinish[searchResult[1]][4] != goodCeneo[2] or ceneoFinish[searchResult[1]][1] != goodCeneo[0]: # or ceneoFinish[searchResult[1]][6] != goodCeneo[3]:
						ceneoFinish[searchResult[1]][4] = goodCeneo[2]
						ceneoFinish[searchResult[1]][1] = goodCeneo[0] # position
						# ceneoFinish[searchResult[1]][6] = goodCeneo[3] # url
						ceneoFinish[searchResult[1]][8] = "Updated"
						updatedGoods += 1
					# del goods2[names2Lower.index(g[4])]
					# del names2Lower[names2Lower.index(g[4])]
					goods2.remove(g)
					names2Lower.remove(g[4])

				if searchResult and searchResult[0] == "onliner":
					goodCeneo = goods[namesLower.index(g[4])]
					# goodCeneo = goods2[names2Lower.index(g)]
					onl = onlinerFinish[searchResult[1]]
					interFinish.append([onl[0], goodCeneo[0], onl[2], onl[3], goodCeneo[2], onl[5], goodCeneo[3], onl[7], "Merged"])
					mergedGoods += 1
					del onlinerFinish[searchResult[1]]
					# del goods2[names2Lower.index(g[4])]
					# del names2Lower[names2Lower.index(g[4])]
					goods2.remove(g)
					names2Lower.remove(g[4])

				if searchResult is None:
					goodCeneo = goods2[names2Lower.index(g[4])]
					ceneoFinish.append(["", goodCeneo[0], goodCeneo[1], "", goodCeneo[2], "", goodCeneo[3], 0, "Added"])
					addedGoods += 1

		elif way == "name":

			for i in inter:
				searchResult = searchInFinish(interFinish, onlinerFinish, ceneoFinish, "name", i)

				if searchResult and searchResult[0] == "inter":
					goodOnliner = goods[namesLower.index(i)]
					goodCeneo = goods2[names2Lower.index(i)]
					if (goodOnliner[2] != interFinish[searchResult[1]][3]) or (goodCeneo[2] != interFinish[searchResult[1]][4]) or (interFinish[searchResult[1]][0] != goodOnliner[0]) or (interFinish[searchResult[1]][5] != goodOnliner[3]) or (interFinish[searchResult[1]][1] != goodCeneo[0]) or (interFinish[searchResult[1]][6] != goodCeneo[3]):
						updatedGoods += sum(list(map(lambda x: 1 if x else 0, [goodOnliner[2] != interFinish[searchResult[1]][3], goodCeneo[2] != interFinish[searchResult[1]][4]])))
						interFinish[searchResult[1]][3] = goodOnliner[2]
						interFinish[searchResult[1]][4] = goodCeneo[2]
						interFinish[searchResult[1]][0] = goodOnliner[0] # position
						interFinish[searchResult[1]][5] = goodOnliner[3] # url
						interFinish[searchResult[1]][1] = goodCeneo[0] # position
						interFinish[searchResult[1]][6] = goodCeneo[3] # url
						interFinish[searchResult[1]][8] = "Updated"
					del goods[namesLower.index(i)]
					del goods2[names2Lower.index(i)]
					# del namesLower[namesLower.index(i)]
					# goods.remove(i)
					namesLower.remove(i)
					# del names2Lower[names2Lower.index(i)]
					# goods2.remove(i)
					names2Lower.remove(i)

				# if searchResult and (searchResult[0] == "onliner" or searchResult[0] == "ceneo"):

				if searchResult and searchResult[0] == "onliner":
					goodCeneo = goods2[names2Lower.index(i)]
					goodOnliner = goods[namesLower.index(i)]
					onl = onlinerFinish[searchResult[1]]
					if onl[3] != goodOnliner[2] or onl[0] != goodOnliner[0] or onl[5] != goodOnliner[3]:
						onlinerFinish[searchResult[1]][3] = goodOnliner[2]
						onlinerFinish[searchResult[1]][0] = goodOnliner[0] # position
						onlinerFinish[searchResult[1]][5] = goodOnliner[3] # url
						updatedGoods += 1
					interFinish.append([onl[0], goodCeneo[0], onl[2], onl[3], goodCeneo[2], onl[5], goodCeneo[3], onl[7], "Merged"])
					mergedGoods += 1
					del onlinerFinish[searchResult[1]]
					del goods[namesLower.index(i)]
					# del namesLower[namesLower.index(i)]
					# goods.remove(i)
					namesLower.remove(i)

				if searchResult and searchResult[0] == "ceneo":
					goodOnliner = goods[namesLower.index(i)]
					goodCeneo = goods2[names2Lower.index(i)]
					cen = ceneoFinish[searchResult[1]]
					if cen[4] != goodCeneo[2] or cen[1] != goodCeneo[0] or cen[6] != goodCeneo[3]:
						ceneoFinish[searchResult[1]][4] = goodCeneo[2]
						ceneoFinish[searchResult[1]][1] = goodCeneo[0] # position
						ceneoFinish[searchResult[1]][6] = goodCeneo[3] # url
						updatedGoods += 1
					interFinish.append([goodOnliner[0], cen[1], goodOnliner[1], goodOnliner[2], cen[4], goodOnliner[3], cen[6], cen[7], "Merged"])
					mergedGoods += 1
					del ceneoFinish[searchResult[1]]
					del goods2[names2Lower.index(i)]
					# del names2Lower[names2Lower.index(i)]
					# goods2.remove(i)
					names2Lower.remove(i)

				if searchResult is None:
					goodOnliner = goods[namesLower.index(i)]
					goodCeneo = goods2[names2Lower.index(i)]
					interFinish.append([goodOnliner[0], goodCeneo[0], goodOnliner[1], goodOnliner[2], goodCeneo[2], goodOnliner[3], goodCeneo[3], 0, "Added"])
					addedGoods += 2
					del goods[namesLower.index(i)]
					del goods2[names2Lower.index(i)]
					# del namesLower[namesLower.index(i)]
					# goods.remove(i)
					namesLower.remove(i)
					# goods2.remove(i)
					names2Lower.remove(i)
					# del names2Lower[names2Lower.index(i)]

			goodsCopy = list(goods)
			for g in goodsCopy: # onliner
				searchResult = searchInFinish(interFinish, onlinerFinish, ceneoFinish, "name", g[1])

				if searchResult and searchResult[0] == "inter":
					goodOnliner = goods[namesLower.index(g[4])]
					if interFinish[searchResult[1]][3] != goodOnliner[2] or interFinish[searchResult[1]][0] != goodOnliner[0] or interFinish[searchResult[1]][5] != goodOnliner[3]:
						interFinish[searchResult[1]][3] = goodOnliner[2] # price
						interFinish[searchResult[1]][0] = goodOnliner[0] # position
						interFinish[searchResult[1]][5] = goodOnliner[3] # url
						interFinish[searchResult[1]][8] = "Updated"
						updatedGoods += 1
					goods.remove(g)
					namesLower.remove(g[4])

				if searchResult and searchResult[0] == "onliner":
					goodOnliner = goods[namesLower.index(g[4])]
					if onlinerFinish[searchResult[1]][3] != goodOnliner[2] or onlinerFinish[searchResult[1]][0] != goodOnliner[0] or onlinerFinish[searchResult[1]][5] != goodOnliner[3]:
						onlinerFinish[searchResult[1]][3] = goodOnliner[2]
						onlinerFinish[searchResult[1]][0] = goodOnliner[0] # position
						onlinerFinish[searchResult[1]][5] = goodOnliner[3] # url
						onlinerFinish[searchResult[1]][8] = "Updated"
						updatedGoods += 1
					# print(goods[namesLower.index(g[4])], end=" ")
					goods.remove(g)
					namesLower.remove(g[4])

				if searchResult and searchResult[0] == "ceneo":
					goodOnliner = goods[namesLower.index(g[4])]
					# goodCeneo = goods2[names2Lower.index(g)]
					cen = ceneoFinish[searchResult[1]]
					interFinish.append([goodOnliner[0], cen[1], goodOnliner[1], goodOnliner[2], cen[4], goodOnliner[3], cen[6], cen[7], "Merged"])
					mergedGoods += 1
					del ceneoFinish[searchResult[1]]
					goods.remove(g)
					namesLower.remove(g[4])

				if searchResult is None:
					goodOnliner = goods[namesLower.index(g[4])]
					onlinerFinish.append([goodOnliner[0], "", goodOnliner[1], goodOnliner[2], "", goodOnliner[3], "", 0, "Added"])
					addedGoods += 1

			goodsCopy = list(goods2)
			for g in goodsCopy: # ceneo
				searchResult = searchInFinish(interFinish, onlinerFinish, ceneoFinish, "name", g[1])

				if searchResult and searchResult[0] == "inter":
					goodCeneo = goods2[names2Lower.index(g[4])]
					if interFinish[searchResult[1]][4] != goodCeneo[2] or interFinish[searchResult[1]][1] != goodCeneo[0] or interFinish[searchResult[1]][6] != goodCeneo[3]:
						interFinish[searchResult[1]][4] = goodCeneo[2]
						interFinish[searchResult[1]][1] = goodCeneo[0] # position
						interFinish[searchResult[1]][6] = goodCeneo[3] # url
						interFinish[searchResult[1]][8] = "Updated"
						updatedGoods += 1
					# del goods2[names2Lower.index(g[4])]
					# del names2Lower[names2Lower.index(g[4])]
					goods2.remove(g)
					names2Lower.remove(g[4])

				if searchResult and searchResult[0] == "ceneo":
					goodCeneo = goods2[names2Lower.index(g[4])]
					if ceneoFinish[searchResult[1]][4] != goodCeneo[2] or ceneoFinish[searchResult[1]][1] != goodCeneo[0] or ceneoFinish[searchResult[1]][6] != goodCeneo[3]:
						ceneoFinish[searchResult[1]][4] = goodCeneo[2]
						ceneoFinish[searchResult[1]][1] = goodCeneo[0] # position
						ceneoFinish[searchResult[1]][6] = goodCeneo[3] # url
						ceneoFinish[searchResult[1]][8] = "Updated"
						updatedGoods += 1
					# del goods2[names2Lower.index(g[4])]
					# del names2Lower[names2Lower.index(g[4])]
					goods2.remove(g)
					names2Lower.remove(g[4])

				if searchResult and searchResult[0] == "onliner":
					goodCeneo = goods[namesLower.index(g[4])]
					# goodCeneo = goods2[names2Lower.index(g)]
					onl = onlinerFinish[searchResult[1]]
					interFinish.append([onl[0], goodCeneo[0], onl[2], onl[3], goodCeneo[2], onl[5], goodCeneo[3], onl[7], "Merged"])
					mergedGoods += 1
					del onlinerFinish[searchResult[1]]
					# del goods2[names2Lower.index(g[4])]
					# del names2Lower[names2Lower.index(g[4])]
					goods2.remove(g)
					names2Lower.remove(g[4])

				if searchResult is None:
					goodCeneo = goods2[names2Lower.index(g[4])]
					ceneoFinish.append(["", goodCeneo[0], goodCeneo[1], "", goodCeneo[2], "", goodCeneo[3], 0, "Added"])
					addedGoods += 1

		print("Added goods: ", addedGoods)
		print("Merged goods: ", mergedGoods)
		print("Updated goods: ", updatedGoods)

		# print(onlinerFinish)
		# print(ceneoFinish)

		onlinerFinish.sort(key=lambda x: x[0])
		ceneoFinish.sort(key=lambda x: x[1])

		wbEnd.save(finishFilename)

		wbEnd = Workbook()
		wsEnd = wbEnd.active

		wsEnd.append(header)

		for item in interFinish:
			wsEnd.append(item)

		for item in onlinerFinish:
			wsEnd.append(item)

		for item in ceneoFinish:
			wsEnd.append(item)

		wbEnd.save(finishFilename)


	else:
		wbEnd = Workbook()
		wsEnd = wbEnd.active
		wsEnd.append(["Number Onliner", "Number Ceneo", "Good's name", "Price on Onliner", "Price on Ceneo", "Link on Onliner", "Link on Ceneo", "Need to parse", "Status"])

		if way == "name":
			for i in inter: # append in table intersections
				# print(namesLower.index(i), names2Lower.index(i))
				goodOnliner = goods[namesLower.index(i)]
				goodCeneo = goods2[names2Lower.index(i)]
				wsEnd.append([goodOnliner[0], goodCeneo[0], goodOnliner[1], goodOnliner[2], goodCeneo[2], goodOnliner[3], goodCeneo[3], 0, "Added"])
				del goods[namesLower.index(i)]
				del namesLower[namesLower.index(i)]
				del goods2[names2Lower.index(i)]
				del names2Lower[names2Lower.index(i)]

		for g in goods: # append in table onliner
			wsEnd.append([g[0], "", g[1], g[2], "", g[3], "", 0, "Added"])

		for g in goods2: #append in table ceneo
			wsEnd.append(["", g[0], g[1], "", g[2], "", g[3], 0, "Added"])

		wbEnd.save(finishFilename)

	deleteDublecates(finishFilename)
	print("Deleted dublecates: " + str(deletedGoods))

def getDB(filename):
	wb = load_workbook(filename)
	ws = wb.active

	return list(list(map(lambda x: x.value, t)) for t in tuple(ws.rows)[1:])

def saveWithHeader(finishFilename, goods, start=2, end=None, header=None):

	if not header:
		wb = load_workbook(finishFilename)
		ws = wb.active

		if not end:
			end = ws.max_row

		if start < end:
			for row in ws[start: end]:
				for cell in row:
					cell.value = None
	else:
		wb = Workbook()
		ws = wb.active

		ws.append(header)

	for ind, row in enumerate(ws[start:start + len(goods) - 1]):
		for ind, i in enumerate(goods[ind]):
			row[ind].value = i

	wb.save(finishFilename)

	return start + len(goods)

def getHeader(filename):
	wb = load_workbook(filename)
	ws = wb.active

	return [t.value for t in tuple(ws.rows)[0]], [t.style for t in tuple(ws.rows)[0]], ws[1]

def getFinishDB(filename):
	wb = load_workbook(filename)
	ws = wb.active

	inter, onliner, ceneo = [], [], []

	for t in tuple(ws.rows)[1:]:
		if t[5].value and "www." in t[5].value:
			t[5].value = t[5].value.replace("www.", "")
		if t[6].value and "www." in t[6].value:
			t[6].value = t[6].value.replace("www.", "")

	for t in tuple(ws.rows)[1:]:
		if t[5].value and t[6].value:
			inter.append(list(map(lambda x: x.value, t)))
		if t[5].value and not t[6].value:
			onliner.append(list(map(lambda x: x.value, t)))
		if not t[5].value and t[6].value:
			ceneo.append(list(map(lambda x: x.value, t)))

	return onliner, ceneo, inter

def searchGood(good, db):

	# import random
	# print("searching..." + random.randint(1, 10))
	# print("good", good[3] in [base[5] for base in db], good[3] in [base[6] for base in db])
	# print(good)
	base5 = [base[5] for base in db]
	base6 = [base[6] for base in db]

	# if good[3] in [base[5] for base in db] and good[3] in [base[6] for base in db]:
	# 	try:
	# 		return ("inter", [base[5] for base in db].index(good[3]))
	# 	except:
	# 		# return None
	# 		pass

	if good[3] in base5 and base6[base5.index(good[3])]:
		try:
			return ("inter", [base[5] for base in db].index(good[3]))
		except:
			# return None
			pass

	if good[3] in base6 and base5[base6.index(good[3])]:
		try:
			return ("inter", [base[6] for base in db].index(good[3]))
		except:
			# return None
			pass

	if good[3] in base5 and not base6[base5.index(good[3])]:
		try:
			return ("onliner", [base[5] for base in db].index(good[3]))
		except:
			# return None
			pass

	if good[3] in base6 and not base5[base6.index(good[3])]:
		try:
			return ("ceneo", [base[6] for base in db].index(good[3]))
		except:
			# return None
			pass

	return None

def needUpdateForGood(newGood, oldGood, way, newOrigin, oldOrigin):
	if newOrigin == "onliner":
		if way == "name":
			fields = [0, 3, 5]
		elif way == "link":
			fields = [0, 2, 3]

	if newOrigin == "ceneo":
		if way == "name":
			fields = [1, 4, 6]
		elif way == "link":
			fields = [1, 2, 4]

	for f in fields:
		if oldGood[f] != newGood[f]:
			# print(f, oldGood, newGood)
			return True

	return False

updatedGoods = 0
mergedGoods = 0
addedGoods = 0

def updateGood(newGood, oldGood, way, newOrigin, oldOrigin):
	if newOrigin == "onliner":
		if way == "name":
			fields = [0, 3, 5]
		elif way == "link":
			fields = [0, 2, 3]

	if newOrigin == "ceneo":
		if way == "name":
			fields = [1, 4, 6]
		elif way == "link":
			fields = [1, 2, 4]

	for f in fields:
		if oldGood[f] != newGood[f]:
			oldGood[f] = newGood[f]

	oldGood[8] = "Updated"

	return oldGood

def addGood(good, site):

	if site == "onliner":
		fields = [0, 2, 3, 5]
	if site == "ceneo":
		fields = [1, 2, 4, 6]

	total = []

	cx = 0
	for r in range(7):
		if r in fields:
			total.append(good[cx])
			cx += 1
		else:
			total.append("")

	total.extend([0, "Added"])

	return total

def saveGoods(inter, onliner, ceneo, filename, header=None):
	from copy import copy, deepcopy

	# new_sheet._styles[cell.get_coordinate()] = copy(
	        # default_sheet._styles[cell.get_coordinate()])
	wb = Workbook()
	ws = wb.active

	if header:
		ws.append(header[0])
		# for ind, t in enumerate(tuple(ws[1])):
		# 	t = copy(header[2][ind]) # header[2] - header of old table
		for ind, t in enumerate(tuple(ws[1])):
			t.style = header[1][ind]
			# t.style = copy(header[1][ind])
	else:
		ws.append(["Number Onliner", "Number Ceneo", "Good's name", "Price on Onliner", "Price on Ceneo", "Link on Onliner", "Link on Ceneo", "Need to parse", "Status"])

	for i in inter:
		ws.append(i)

	for o in sorted(onliner, key=lambda x: x[0]):
		ws.append(o)

	for c in sorted(ceneo, key=lambda x: x[1]):
		ws.append(c)


	wb.save(filename)

def intersectionAndSave(onlinerFilename, ceneoFilename, finishFilename):
		
	try:
		onlinerDatebase = getDB(onlinerFilename)
	except:
		onlinerDatebase = []

	try:
		ceneoDatebase = getDB(ceneoFilename)
	except:
		ceneoDatebase = []

	global updatedGoods
	global mergedGoods
	global addedGoods

	print(onlinerDatebase[:5])
	print(ceneoDatebase[:5])

	if os.path.isfile(finishFilename):

		deleteDublecates(finishFilename)
		print("Deleted before intersection:", deletedGoods, "goods")

		onlinerFinish, ceneoFinish, interFinish = getFinishDB(finishFilename)

		for onl in onlinerDatebase:
			# print(onl)
			searchResult = searchGood(onl, interFinish)

			# print(searchResult)

			if searchResult and (searchResult[0] == "inter" or searchResult[0] == "onliner"):
				if needUpdateForGood(addGood(onl, "onliner"), interFinish[searchResult[1]], "link", "onliner", "inter"):
					interFinish[searchResult[1]] = updateGood(addGood(onl, "onliner"), interFinish[searchResult[1]], "link", "onliner", "inter")
					updatedGoods += 1
				elif needUpdateForGood(addGood(onl, "onliner"), interFinish[searchResult[1]], "name", "onliner", "inter"):
					interFinish[searchResult[1]] = updateGood(addGood(onl, "onliner"), interFinish[searchResult[1]], "name", "onliner", "inter")
					updatedGoods += 1

			else:
				searchResult = searchGood(onl, onlinerFinish)

				if searchResult and searchResult[0] == "onliner":
					if needUpdateForGood(addGood(onl, "onliner"), onlinerFinish[searchResult[1]], "link", "onliner", "onliner"):
						onlinerFinish[searchResult[1]] = updateGood(addGood(onl, "onliner"), onlinerFinish[searchResult[1]], "link", "onliner", "onliner")
						updatedGoods += 1
					elif needUpdateForGood(addGood(onl, "onliner"), onlinerFinish[searchResult[1]], "name", "onliner", "onliner"):
						onlinerFinish[searchResult[1]] = updateGood(addGood(onl, "onliner"), onlinerFinish[searchResult[1]], "name", "onliner", "onliner")
						updatedGoods += 1

				elif searchResult is None:
					# print(addGood(onl, "onliner"), addGood(onl, "onliner")[5] in [o[3] for o in onlinerDatebase])
					onlinerFinish.append(addGood(onl, "onliner"))
					addedGoods += 1

		for cen in ceneoDatebase:
			searchResult = searchGood(cen, interFinish)
			# try:
			# 	if cen[0] == 37:
			# 		print(cen, searchResult)
			# except:
			# 	pass

			if searchResult and (searchResult[0] == "inter" or searchResult[0] == "ceneo"):
				if needUpdateForGood(addGood(cen, "ceneo"), interFinish[searchResult[1]], "link", "ceneo", "inter"):
					interFinish[searchResult[1]] = updateGood(addGood(cen, "ceneo"), interFinish[searchResult[1]], "link", "ceneo", "inter")
					updatedGoods += 1
				elif needUpdateForGood(addGood(cen, "ceneo"), interFinish[searchResult[1]], "name", "ceneo", "inter"):
					interFinish[searchResult[1]] = updateGood(addGood(cen, "ceneo"), interFinish[searchResult[1]], "name", "ceneo", "inter")
					updatedGoods += 1

			# elif searchResult and searchResult[0] == "ceneo":
			# 	if needUpdateForGood(addGood(cen, "ceneo"), interFinish[searchResult[1]], "link", "ceneo", "inter"):
			# 		interFinish[searchResult[1]] = updateGood(addGood(cen, "ceneo"), interFinish[searchResult[1]], "link", "ceneo", "inter")
			# 		updatedGoods += 1
			# 	elif needUpdateForGood(addGood(cen, "ceneo"), interFinish[searchResult[1]], "name", "ceneo", "inter"):
			# 		interFinish[searchResult[1]] = updateGood(addGood(cen, "ceneo"), interFinish[searchResult[1]], "name", "ceneo", "inter")
			# 		updatedGoods += 1

			else:
				searchResult = searchGood(cen, ceneoFinish)

				if searchResult and searchResult[0] == "ceneo":
					if needUpdateForGood(addGood(cen, "ceneo"), ceneoFinish[searchResult[1]], "link", "ceneo", "ceneo"):
						ceneoFinish[searchResult[1]] = updateGood(addGood(cen, "ceneo"), ceneoFinish[searchResult[1]], "link", "ceneo", "ceneo")
						# if cen[0] == 37:
						# 	print(ceneoFinish[searchResult[1]])
						updatedGoods += 1
					elif needUpdateForGood(addGood(cen, "ceneo"), ceneoFinish[searchResult[1]], "name", "ceneo", "ceneo"):
						ceneoFinish[searchResult[1]] = updateGood(addGood(cen, "ceneo"), ceneoFinish[searchResult[1]], "name", "ceneo", "ceneo")
						# if cen[0] == 37:
						# 	print(ceneoFinish[searchResult[1]])
						updatedGoods += 1

				elif searchResult is None:
					# print(addGood(cen, "ceneo"))
					ceneoFinish.append(addGood(cen, "ceneo"))
					addedGoods += 1

		# saveGoods(interFinish, onlinerFinish, ceneoFinish, finishFilename, header=getHeader(finishFilename))
		last = saveWithHeader(finishFilename, interFinish)
		last = saveWithHeader(finishFilename, sorted(onlinerFinish, key=lambda x: x[0]), last)
		saveWithHeader(finishFilename, sorted(ceneoFinish, key=lambda x: x[1]), last)

		deleteDublecates(finishFilename)
		print("Deleted after intersection:", deletedGoods, "goods")

	else:

		onlinerInd = [[good[1].lower(), ind] for ind, good in enumerate(onlinerDatebase)]
		ceneoInd = [[good[1].lower(), ind] for ind, good in enumerate(ceneoDatebase)]

		interInd = []
		inter = []

		temp = [c[0] for c in ceneoInd]
		for good, ind in onlinerInd:
			if good in temp:
				# interInd.append([good, ind, temp.index(good)])
				inter.append([onlinerDatebase[ind], ceneoDatebase[temp.index(good)]])
				onlinerDatebase[ind] = None
				ceneoDatebase[temp.index(good)] = None

		# temp = [c[0] for c in onlinerInd]
		# for good, ind in ceneoInd:
		# 	if good in temp:
		# 		# interInd.append([good, ind, temp.index(good)])
		# 		inter.append([onlinerDatebase[temp.index(good)], ceneoDatebase[ind]])
		# 		onlinerDatebase[temp.index(good)] = None
		# 		ceneoDatebase[ind] = None


		# for i in interInd:
		# 	if onlinerDatebase[i] and ceneoDatabase[i]:
		# 		inter.append(onlinerDatebase[i[1]], ceneoDatebase[i[2]])

		# print("inter", inter)


		for i in range(len(inter)):
			# print(addGood(inter[i][0], "onliner"), addGood(inter[i][1], "ceneo"), end=" ")
			inter[i] = updateGood(addGood(inter[i][0], "onliner"), addGood(inter[i][1], "ceneo"), "name", "onliner", "ceneo")
			mergedGoods += 1
			# print(inter[i])

		onliner, ceneo = [], []

		for i in range(len(onlinerDatebase)):
			if onlinerDatebase[i]:
				onliner.append(addGood(onlinerDatebase[i], "onliner"))

		for i in range(len(ceneoDatebase)):
			if ceneoDatebase[i]:
				ceneo.append(addGood(ceneoDatebase[i], "ceneo"))

		# print(inter)

		# saveGoods(inter, onliner, ceneo, finishFilename)
		last = saveWithHeader(finishFilename, inter, header=["Number Onliner", "Number Ceneo", "Good's name", "Price on Onliner", "Price on Ceneo", "Link on Onliner", "Link on Ceneo", "Need to parse", "Status"])
		last = saveWithHeader(finishFilename, sorted(onliner, key=lambda x: x[0]), last)
		saveWithHeader(finishFilename, sorted(ceneo, key=lambda x: x[1]), last)

	print("Added goods:", addedGoods)
	print("Updated goods:", updatedGoods)
	print("Merged goods:", mergedGoods)

def updatePricesWithoutThreadings(sheetFilename, curr, time=None):

	print("Opening file: " + sheetFilename)

	wb = load_workbook(sheetFilename)
	sheet = wb.active

	# for s in sheet["H"]:
	# 	print(str(s.value) + " ", end="")

	needUpdate = list(inx for inx, s in enumerate(sheet["H"]) if s.value == 1)
	# print(needUpdate)

	columnsToUpdate = list(tuple(sheet.rows)[n] for n in needUpdate)

	linksAddress = (5, 6)
	pricesAddress = (3, 4)

	if len(columnsToUpdate) > 0:
		print("Need to update: " + str(len(columnsToUpdate)) + " good(s)")
	else:
		print("Update is not necessary")

	for c in range(len(columnsToUpdate)):

		sec = datetime.datetime.now()

		urlOnliner = columnsToUpdate[c][linksAddress[0]].value

		if urlOnliner:
			try:
				html = requests.get(urlOnliner, verify=True, timeout=time).content

				soup = BeautifulSoup(html, "html.parser")

				try:
					onlinerPrice = round(float(soup.find("span", "product-aside__price--primary").text.strip().replace(",", ".")[:-2]) / curr["BYN"], 2)
				except:
					onlinerPrice = "Нет в наличии"

				columnsToUpdate[c][8].value = "Updated"
				columnsToUpdate[c][7].value = 1 # put 0 to prevent the product from being updated
				columnsToUpdate[c][pricesAddress[0]].value = onlinerPrice

			except:
				# onlinerPrice = "Ошибка получения"
				columnsToUpdate[c][8].value = "Error"
				columnsToUpdate[c][7].value = 1
				return False

		urlCeneo = columnsToUpdate[c][linksAddress[1]].value

		if urlCeneo:
			try:
				html = requests.get(urlCeneo, verify=True, timeout=time).content

				soup = BeautifulSoup(html, "html.parser")

				try:
					ceneoPrice = soup.find("span", {"class": "price"})
					ceneoPrice = round(float(ceneoPrice.find("span", "value").text + ceneoPrice.find("span", "penny").text.replace(",", ".")) / curr["PLN"], 2)
				except:
					ceneoPrice = "Нет в наличии"

				columnsToUpdate[c][8].value = "Updated"
				columnsToUpdate[c][7].value = 1 # put 0 to prevent the product from being updated
				columnsToUpdate[c][pricesAddress[1]].value = ceneoPrice

			except:
				# ceneoPrice = "Ошибка получения"
				columnsToUpdate[c][8].value = "Error"
				columnsToUpdate[c][7].value = 1
				return False

		print(str(c + 1) + " / " + str(len(columnsToUpdate)) + " good updated! Spent time: " + str(datetime.datetime.now() - sec))

	# print(onlinerPrice, ceneoPrice)

	print("Saving file: " + sheetFilename)

	wb.save(sheetFilename)

def oneThread(columnsToUpdate, c, curr, time=None, proxys=None):

	linksAddress = (5, 6)
	pricesAddress = (3, 4)

	sec = datetime.datetime.now()

	urlOnliner = columnsToUpdate[c][linksAddress[0]].value

	if urlOnliner:
		try:
			html = requests.get(urlOnliner, timeout=time, proxies=proxys).content

			soup = BeautifulSoup(html, "html.parser")

			try:
				onlinerPrice = round(float(soup.find("span", "product-aside__price--primary").text.strip().replace(",", ".")[:-2]) / curr["BYN"], 2)
			except:
				onlinerPrice = "Нет в наличии"

			columnsToUpdate[c][8].value = "Updated"
			columnsToUpdate[c][7].value = 1 # put 0 to prevent the product from being updated
			columnsToUpdate[c][pricesAddress[0]].value = onlinerPrice

		except:
			# onlinerPrice = "Ошибка получения"
			columnsToUpdate[c][8].value = "Error"
			columnsToUpdate[c][7].value = 1
			return False

	urlCeneo = columnsToUpdate[c][linksAddress[1]].value

	if urlCeneo:
		try:
			html = requests.get(urlCeneo, timeout=time, proxies=proxys).content

			soup = BeautifulSoup(html, "html.parser")

			try:
				ceneoPrice = soup.find("span", {"class": "price"})
				ceneoPrice = round(float(ceneoPrice.find("span", "value").text + ceneoPrice.find("span", "penny").text.replace(",", ".")) / curr["PLN"], 2)
			except:
				ceneoPrice = "Нет в наличии"

			columnsToUpdate[c][8].value = "Updated"
			columnsToUpdate[c][7].value = 1 # put 0 to prevent the product from being updated
			columnsToUpdate[c][pricesAddress[1]].value = ceneoPrice

		except:
			# ceneoPrice = "Ошибка получения"
			columnsToUpdate[c][8].value = "Error"
			columnsToUpdate[c][7].value = 1
			return False

	print(str(c + 1) + " / " + str(len(columnsToUpdate)) + " good updated! Spent time: " + str(datetime.datetime.now() - sec))

def updatePrices(sheetFilename, curr, time=None, proxys=None):

	from multiprocessing.dummy import Pool as ThreadPool 
	pool = ThreadPool(50)

	print("Opening file: " + sheetFilename)

	wb = load_workbook(sheetFilename)
	sheet = wb.active

	# for s in sheet["H"]:
	# 	print(str(s.value) + " ", end="")

	needUpdate = list(inx for inx, s in enumerate(sheet["H"]) if s.value == 1)
	# print(needUpdate)

	columnsToUpdate = list(tuple(sheet.rows)[n] for n in needUpdate)

	linksAddress = (5, 6)
	pricesAddress = (3, 4)

	if len(columnsToUpdate) > 0:
		print("Need to update: " + str(len(columnsToUpdate)) + " good(s)")
	else:
		print("Update is not necessary")

	# for c in range(len(columnsToUpdate)):
	pool.starmap(oneThread, zip([columnsToUpdate] * len(columnsToUpdate), range(len(columnsToUpdate)), [curr] * len(columnsToUpdate), [time] * len(columnsToUpdate), [proxys] * len(columnsToUpdate)))

	print("Saving file: " + sheetFilename)

	wb.save(sheetFilename)

def getCurrencies(sheetFilename, api_key, base="USD"):
	wb = load_workbook(sheetFilename)
	ws = wb.active

	needUpdate = list(tuple(ws.rows)[i] for i, w in enumerate(ws["C"][1:], 1) if w.value == 1)
	tableCurr = list(tuple(ws.rows)[1:])

	# print(needUpdate)
	USDPLN, USDBYN = 0, 0

	if needUpdate:
		curr = list(n[0].value for n in needUpdate)
		print("Need update this currencies: " + ", ".join(curr))
		currencies = ",".join(list(n[0].value.upper() for n in needUpdate))
		# api_key = str(api_key.encode("ascii", "ignore"))
		json = requests.get("http://apilayer.net/api/live?access_key=" + api_key + "&currencies=" + currencies + "&source=" + base + "&format=1").json()
		quotes = json["quotes"]
		for n in needUpdate:
			n[1].value = quotes["USD" + n[0].value.upper()]
			n[2].value = 0
			print("USD" + n[0].value.upper() + ": " + str(n[1].value))

	else:
		print("Currency update isn't needed")

	for t in tableCurr:
		if t[0].value == "PLN":
			USDPLN = t[1].value
		if t[0].value == "BYN":
			USDBYN = t[1].value

	wb.save(sheetFilename)
	print("BYN = " + str(USDBYN))
	print("PLN = " + str(USDPLN))
	return {"PLN": USDPLN, "BYN": USDBYN}

def runTask(mainFilename, multiprocessing=False):
	wb = load_workbook(mainFilename)
	ws = wb.active

	priceParse = [] # filenames
	makeData = []  # onliner url, ceneo url, filename

	# print(ws["C"][1])
	taskForParse = list([w, ws["C"][w].value] for w in range(ws.max_row) if ws["C"][w].value != None)[1:]
	# print(taskForParse)
	for column, task in taskForParse:
		if task == 2:
			priceParse.append(ws["D"][column].value)
		if task == 1:
			makeData.append([ws["A"][column].value, ws["B"][column].value, ws["D"][column].value])

	# print(priceParse, makeData)

	file = open("api_key.txt", "r")
	apikey = file.read().strip()
	file.close()

	curr = getCurrencies("currencies.xlsx", apikey)

	for p in priceParse:
		try:
			if multiprocessing:
				updatePrices(p, curr)
			else:
				updatePricesWithoutThreadings(p, curr)
		except:
			print("Файл {} не существует или произошла ошибка при парсинге".format(p))

	for m in makeData:
		filenameOnliner, filenameCeneo = "", ""
		if m[0]:
			filenameOnliner = m[0][len(m[0]) - 1 - m[0][::-1].index("/") + 1:] + "(onliner.by).xlsx"
			searchGoods(m[0], "onliner", curr, filenameOnliner)
		if m[1]:
			filenameCeneo = m[1][len(m[1]) - 1 - m[1][::-1].index("/") + 1:] + "(ceneo.pl).xlsx"
			searchGoods(m[1], "ceneo", curr, filenameCeneo)
		# IntersectionAndSave(filenameOnliner, filenameCeneo, m[2], "link")
		# IntersectionAndSave(filenameOnliner, filenameCeneo, m[2], "name")
		intersectionAndSave(filenameOnliner, filenameCeneo, m[2])


# print(getAllGoodsFromPageCeneo(page)[1])
# page = "https://catalog.onliner.by/coffee"
# searchGoods(page, "onliner", "onliner.by")
# page = "https://www.ceneo.pl/Ekspresy_do_kawy"
# searchGoods(page, "ceneo", "ceneo.pl")

# IntersectionAndSave("onliner.by", "ceneo.pl", "finishTable")

# updatePricesWithoutThreadings("finishTable")
# proxys = {
# 		"https": "https://139.59.125.12:8080",
# 		"http": "http://89.187.217.115:80"
# 	}
# updatePrices("finishTable", time=10)

if __name__ == "__main__":
	totalTime = datetime.datetime.now()
	runTask("main.xlsx")
	# intersectionAndSave("coffee(onliner.by).xlsx", "Ekspresy_do_kawy(ceneo.pl).xlsx", "coffee.xlsx")
	# saveWithHeader("coffee.xlsx", [[1, 213, 234], [2, 132, 123]], 5)
	# print(updateGood([5, "", "naming", 123, "", "https", "", 0, "Added"], ["", 10, "naming", "", 234, "", "http", 0, "Added"], "name", "ceneo", "ceneo"))
	# deleteDublecates("coffee.xlsx")
	# print(deletedGoods)
	# IntersectionAndSave("notebook(onliner.by).xlsx", "Laptopy(ceneo.pl).xlsx", "notebooks.xlsx")
	print("Totally spent time: " + str(datetime.datetime.now() - totalTime))